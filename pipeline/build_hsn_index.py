"""
pipeline/build_hsn_index.py
============================
ONE-TIME OFFLINE BUILD SCRIPT.
Run once per machine after cloning the project.

What it does
------------
1. Parses HSN_SAC.xlsx (official e-invoice master, both sheets):
   - HSN_MSTR: 14,709 leaf-level 8-digit codes
   - SAC_MSTR:    649 leaf-level 6-digit codes
2. Builds a full hierarchical description for each leaf by walking parent
   chapter → heading → sub-heading rows (leaf text alone is often just
   "OTHER" or a single word — the hierarchy gives semantic context).
3. Joins every code with its correct GST rate via gst_rate_schedule.py.
4. Guesses a standard unit of measure from description keywords.
5. Saves JSON masters + TF-IDF search indices to pipeline/hsn_data/.

Why TF-IDF (not sentence-transformers)
---------------------------------------
HSN/SAC descriptions are dense legal/technical jargon where exact term
overlap is the strongest match signal. TF-IDF rewards that directly,
needs no model download, runs fully offline, uses ~50MB RAM, and takes
<5 seconds to build. The full 23k-vector search query takes <10ms on CPU.

Usage
-----
    cd <project_root>
    python pipeline/build_hsn_index.py --xlsx /path/to/HSN_SAC.xlsx

Outputs (commit these — never need to re-run unless you refresh the xlsx)
--------------------------------------------------------------------------
    pipeline/hsn_data/hsn_master.json
    pipeline/hsn_data/sac_master.json
    pipeline/hsn_data/hsn_vectorizer.joblib
    pipeline/hsn_data/hsn_matrix.joblib
    pipeline/hsn_data/hsn_meta.json
    pipeline/hsn_data/sac_vectorizer.joblib
    pipeline/hsn_data/sac_matrix.joblib
    pipeline/hsn_data/sac_meta.json
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path

HERE = Path(__file__).parent
sys.path.insert(0, str(HERE))

import openpyxl
from gst_rate_schedule import lookup_rate, guess_unit

DATA_DIR = HERE / "hsn_data"


# ─────────────────────────────────────────────────────────────────────────────
#  Text helpers
# ─────────────────────────────────────────────────────────────────────────────

def _clean(s) -> str:
    if s is None:
        return ""
    s = str(s)
    s = s.replace("_x000D_", " ").replace("~", " — ")
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _dedup_join(parts: list[str]) -> str:
    """Join hierarchy parts, dropping consecutive exact duplicates and bare punctuation."""
    out: list[str] = []
    for p in parts:
        p = p.strip(" :;,.")
        if not p:
            continue
        if out and out[-1].lower() == p.lower():
            continue
        out.append(p)
    return " — ".join(out)


# ─────────────────────────────────────────────────────────────────────────────
#  Parsers
# ─────────────────────────────────────────────────────────────────────────────

def parse_hsn(xlsx_path: Path) -> list[dict]:
    """
    The sheet is ordered hierarchically: chapter(2) → heading(4) →
    subheading(6) → tariff-item(8). Parent rows always precede children.
    We track a running parent stack and build a full description for each
    leaf by concatenating chapter→heading→subheading→leaf text.
    """
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb["HSN_MSTR"]

    records: list[dict] = []
    seen: set[str] = set()
    parent: dict[int, str] = {2: "", 4: "", 6: ""}

    for row in ws.iter_rows(min_row=2, values_only=True):
        code_raw, desc_raw = row[0], row[1]
        if code_raw is None:
            continue
        code_str = str(code_raw).strip().replace(" ", "")
        if not code_str.isdigit():
            continue
        desc = _clean(desc_raw)
        length = len(code_str)

        if length in parent:
            parent[length] = desc
            continue          # hierarchy row — context only, not a leaf

        if length != 8:
            continue          # rare malformed-length rows

        if code_str in seen:
            continue
        seen.add(code_str)

        full_desc = _dedup_join([parent[2], parent[4], parent[6], desc])
        rate = lookup_rate(code_str)
        unit = guess_unit(full_desc)

        records.append({
            "code": code_str,
            "description": full_desc,
            "leaf_description": desc,
            "gst_rate": rate,
            "unit": unit,
        })

    wb.close()
    return records


def parse_sac(xlsx_path: Path) -> list[dict]:
    """SAC hierarchy: 2 → 4 → 6 digits. Leaf = 6-digit."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb["SAC_MSTR"]

    records: list[dict] = []
    seen: set[str] = set()
    parent: dict[int, str] = {2: "", 4: ""}

    for row in ws.iter_rows(min_row=2, values_only=True):
        code_raw, desc_raw = row[0], row[1]
        if code_raw is None:
            continue
        code_str = str(code_raw).strip().replace(" ", "")
        if not code_str.isdigit():
            continue
        desc = _clean(desc_raw)
        length = len(code_str)

        if length in parent:
            parent[length] = desc
            continue

        if length != 6:
            continue

        if code_str in seen:
            continue
        seen.add(code_str)

        full_desc = _dedup_join([parent[2], parent[4], desc])
        rate = lookup_rate(code_str, is_service=True)

        records.append({
            "code": code_str,
            "description": full_desc,
            "leaf_description": desc,
            "gst_rate": rate,
        })

    wb.close()
    return records


# ─────────────────────────────────────────────────────────────────────────────
#  TF-IDF index builder
# ─────────────────────────────────────────────────────────────────────────────

def build_index(records: list[dict], prefix: str) -> None:
    """
    Fit a TF-IDF vectorizer over all descriptions and save:
      {prefix}_vectorizer.joblib  — fitted vectorizer
      {prefix}_matrix.joblib      — sparse TF-IDF matrix (n_docs × vocab)
      {prefix}_meta.json          — parallel list of records (same order)
    """
    import joblib
    from sklearn.feature_extraction.text import TfidfVectorizer

    texts = [r["description"] for r in records]
    print(f"  Fitting TF-IDF over {len(texts)} descriptions …")

    vec = TfidfVectorizer(
        lowercase=True,
        ngram_range=(1, 2),   # unigrams + bigrams for multi-word product terms
        sublinear_tf=True,    # log(1+tf) dampens very frequent terms
        max_df=0.60,          # drop words in >60% of docs (pure boilerplate)
        min_df=1,
    )
    matrix = vec.fit_transform(texts)

    DATA_DIR.mkdir(parents=True, exist_ok=True)
    joblib.dump(vec,    DATA_DIR / f"{prefix}_vectorizer.joblib")
    joblib.dump(matrix, DATA_DIR / f"{prefix}_matrix.joblib")
    with open(DATA_DIR / f"{prefix}_meta.json", "w", encoding="utf-8") as f:
        json.dump(records, f, ensure_ascii=False)

    print(f"  Saved {prefix}_vectorizer.joblib + {prefix}_matrix.joblib "
          f"({matrix.shape[0]} docs × {matrix.shape[1]} vocab terms) + {prefix}_meta.json")


# ─────────────────────────────────────────────────────────────────────────────
#  Entry point
# ─────────────────────────────────────────────────────────────────────────────

def main() -> None:
    ap = argparse.ArgumentParser(
        description="Build HSN/SAC TF-IDF search index from official xlsx."
    )
    ap.add_argument("--xlsx", required=True, help="Path to HSN_SAC.xlsx")
    ap.add_argument(
        "--skip-index", action="store_true",
        help="Only produce JSON masters, skip TF-IDF build (for inspection)"
    )
    args = ap.parse_args()

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        print(f"ERROR: {xlsx_path} not found", file=sys.stderr)
        sys.exit(1)

    DATA_DIR.mkdir(parents=True, exist_ok=True)

    print("Parsing HSN_MSTR (goods) …")
    hsn = parse_hsn(xlsx_path)
    print(f"  {len(hsn)} leaf-level 8-digit HSN codes")
    with open(DATA_DIR / "hsn_master.json", "w", encoding="utf-8") as f:
        json.dump(hsn, f, ensure_ascii=False, indent=1)

    print("Parsing SAC_MSTR (services) …")
    sac = parse_sac(xlsx_path)
    print(f"  {len(sac)} leaf-level 6-digit SAC codes")
    with open(DATA_DIR / "sac_master.json", "w", encoding="utf-8") as f:
        json.dump(sac, f, ensure_ascii=False, indent=1)

    if args.skip_index:
        print("Skipping TF-IDF index build (--skip-index).")
        return

    print("Building HSN search index …")
    build_index(hsn, "hsn")
    print("Building SAC search index …")
    build_index(sac, "sac")

    print(
        "\nDone. pipeline/hsn_data/ is ready.\n"
        "Commit it to your repo — this never needs to re-run unless you\n"
        "refresh HSN_SAC.xlsx from the GST portal.\n"
        "\nNext step: python test_pipeline.py"
    )


if __name__ == "__main__":
    main()